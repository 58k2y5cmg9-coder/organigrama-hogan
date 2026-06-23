"""
Analizador de Hojas de Vida — powered by Groq
Extrae: datos personales + experiencia laboral completa
Exporta a Excel con 2 pestañas
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading, json, traceback
import pdfplumber, docx, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from groq import Groq

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

C = {
    "bg":    "#0D0D1A", "card":  "#141428", "input": "#1A1A35",
    "ac":    "#F97316", "ac2":   "#FED7AA", "ok":    "#10B981",
    "err":   "#EF4444", "warn":  "#F59E0B", "txt":   "#E2E8F0",
    "dim":   "#94A3B8", "borde": "#2D2D5E",
    "lok":   "#34D399", "lerr":  "#F87171", "linfo": "#60A5FA", "lwarn": "#FBBF24",
}
MODELO = "llama-3.3-70b-versatile"


# ─── EXTRACCIÓN DE TEXTO ──────────────────────────────────────────────────────
def extraer_texto(ruta):
    ext = Path(ruta).suffix.lower()
    if ext == ".pdf":
        texto = ""
        with pdfplumber.open(ruta) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t: texto += t + "\n"
        return texto
    elif ext in (".docx", ".doc"):
        d = docx.Document(ruta)
        return "\n".join(p.text for p in d.paragraphs if p.text.strip())
    elif ext == ".txt":
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    return ""


# ─── ANÁLISIS CON GROQ ────────────────────────────────────────────────────────
def analizar_con_groq(texto_cv, api_key):
    cliente = Groq(api_key=api_key)

    system_prompt = (
        "Eres un extractor experto de datos de hojas de vida. "
        "SIEMPRE respondes con JSON puro y válido. "
        "Nunca agregas markdown, explicaciones ni texto fuera del JSON. "
        "Para experiencia laboral, extrae TODOS los trabajos que encuentres. "
        "Si un trabajo está actualmente en curso, usa exactamente la cadena 'En curso' en fecha_fin. "
        "Para las fechas usa el formato que aparece en el CV (ej: 'Enero 2020', '03/2021', '2019'). "
        "Si no encuentras un dato usa null."
    )

    prompt = (
        "Extrae la siguiente información de esta hoja de vida.\n\n"
        "Hoja de vida:\n---\n"
        + texto_cv[:7000]
        + "\n---\n\n"
        "Devuelve EXACTAMENTE este JSON con los datos encontrados:\n"
        "{\n"
        '  "nombre_completo": null,\n'
        '  "correo": null,\n'
        '  "celular": null,\n'
        '  "universidad": null,\n'
        '  "nivel_estudios": null,\n'
        '  "titulo": null,\n'
        '  "anio_graduacion": null,\n'
        '  "experiencia_laboral": [\n'
        '    {\n'
        '      "empresa": "nombre de la empresa",\n'
        '      "cargo": "cargo o puesto desempeñado",\n'
        '      "fecha_inicio": "mes y año de inicio",\n'
        '      "fecha_fin": "mes y año de fin o En curso"\n'
        '    }\n'
        '  ]\n'
        "}\n\n"
        "IMPORTANTE: experiencia_laboral debe ser una lista con TODOS los trabajos. "
        "Si no hay experiencia laboral, usa una lista vacía []."
    )

    respuesta = cliente.chat.completions.create(
        model=MODELO,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": prompt}
        ],
        temperature=0,
        max_tokens=2000,
    )

    texto = respuesta.choices[0].message.content.strip()
    # Limpiar markdown si aparece
    if "```" in texto:
        for parte in texto.split("```"):
            limpio = parte.strip().lstrip("json").strip()
            if limpio.startswith("{"): texto = limpio; break
    # Asegurar que experiencia_laboral sea lista
    datos = json.loads(texto)
    if not isinstance(datos.get("experiencia_laboral"), list):
        datos["experiencia_laboral"] = []
    return datos


# ─── EXPORTAR EXCEL ───────────────────────────────────────────────────────────
def _borde(color="2D2D5E"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _celda(ws, fila, col, valor, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=fila, column=col, value=valor)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c

def exportar_excel(datos, ruta):
    wb = openpyxl.Workbook()

    # ── HOJA 1: Datos Personales ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Datos Personales"

    H1_FILL   = PatternFill("solid", fgColor="431407")
    ALT_FILL  = PatternFill("solid", fgColor="1C1917")
    NORM_FILL = PatternFill("solid", fgColor="0C0A09")
    H1_FONT   = Font(name="Calibri", bold=True, color="FED7AA", size=11)
    DATA_FONT = Font(name="Calibri", color="E2E8F0", size=10)
    DIM_FONT  = Font(name="Calibri", color="94A3B8", size=9, italic=True)
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    BRD = _borde("7C2D12")

    cols1 = [
        ("N°",8),("Nombre Completo",30),("Correo",32),("Celular",18),
        ("Universidad",28),("Nivel Estudios",16),("Título / Carrera",26),
        ("Año Grad.",12),("Experiencias",14),("Archivo Origen",30),
    ]
    for ci,(titulo,ancho) in enumerate(cols1,1):
        _celda(ws1,1,ci,titulo, font=H1_FONT, fill=H1_FILL, align=AC, border=BRD)
        ws1.column_dimensions[get_column_letter(ci)].width = ancho
    ws1.row_dimensions[1].height = 28

    for i,fila in enumerate(datos,1):
        fill = ALT_FILL if i%2==0 else NORM_FILL
        exp  = fila.get("experiencia_laboral") or []
        vals = [
            i,
            fila.get("nombre_completo") or "No encontrado",
            fila.get("correo")          or "No encontrado",
            fila.get("celular")         or "No encontrado",
            fila.get("universidad")     or "No encontrado",
            fila.get("nivel_estudios")  or "No encontrado",
            fila.get("titulo")          or "No encontrado",
            fila.get("anio_graduacion") or "No encontrado",
            len(exp),
            fila.get("_archivo",""),
        ]
        for ci,v in enumerate(vals,1):
            fnt = DIM_FONT if ci==len(vals) else DATA_FONT
            aln = AC if ci in (1,9) else AL
            _celda(ws1,i+1,ci,v, font=fnt, fill=fill, align=aln, border=BRD)
        ws1.row_dimensions[i+1].height = 22

    ft=len(datos)+2
    c=ws1.cell(row=ft,column=1,value=f"Total: {len(datos)} personas")
    c.font=Font(name="Calibri",bold=True,color="FED7AA",size=10)
    c.fill=PatternFill("solid",fgColor="1C0A00")
    ws1.merge_cells(f"A{ft}:{get_column_letter(len(cols1))}{ft}")
    c.alignment=AC
    ws1.freeze_panes="A2"

    # ── HOJA 2: Experiencia Laboral ───────────────────────────────────────────
    ws2 = wb.create_sheet(title="Experiencia Laboral")

    H2_FILL  = PatternFill("solid", fgColor="1A3A1A")
    ALT2     = PatternFill("solid", fgColor="0F1F0F")
    NORM2    = PatternFill("solid", fgColor="080F08")
    EC_FILL  = PatternFill("solid", fgColor="14350C")   # En curso → verde oscuro
    H2_FONT  = Font(name="Calibri", bold=True, color="86EFAC", size=11)
    EC_FONT  = Font(name="Calibri", bold=True, color="4ADE80", size=10)
    BRD2     = _borde("166534")

    cols2 = [
        ("N° Persona",12),("Nombre",28),("Empresa",32),("Cargo",28),
        ("Fecha Inicio",16),("Fecha Fin",16),("Estado",12),
    ]
    for ci,(titulo,ancho) in enumerate(cols2,1):
        _celda(ws2,1,ci,titulo, font=H2_FONT, fill=H2_FILL, align=AC, border=BRD2)
        ws2.column_dimensions[get_column_letter(ci)].width = ancho
    ws2.row_dimensions[1].height = 28

    fila_exp = 2
    for i,persona in enumerate(datos,1):
        nombre  = persona.get("nombre_completo") or "Sin nombre"
        exps    = persona.get("experiencia_laboral") or []

        if not exps:
            fill = ALT2 if fila_exp%2==0 else NORM2
            for ci,v in enumerate([i, nombre, "Sin experiencia registrada","","","",""],1):
                _celda(ws2,fila_exp,ci,v,
                       font=Font(name="Calibri",color="94A3B8",size=10,italic=True),
                       fill=fill, align=AL if ci>1 else AC, border=BRD2)
            ws2.row_dimensions[fila_exp].height = 22
            fila_exp+=1
            continue

        for exp in exps:
            en_curso = str(exp.get("fecha_fin","")).strip().lower() in (
                "en curso","actual","presente","current","actualidad","a la fecha","")
            fecha_fin = "En curso" if en_curso else (exp.get("fecha_fin") or "No especificada")
            estado    = "🟢 En curso" if en_curso else "✓ Finalizado"

            fill  = EC_FILL   if en_curso else (ALT2 if fila_exp%2==0 else NORM2)
            f_fin = EC_FONT   if en_curso else DATA_FONT
            f_est = EC_FONT   if en_curso else Font(name="Calibri",color="94A3B8",size=10)

            vals2 = [i, nombre,
                     exp.get("empresa")      or "No especificada",
                     exp.get("cargo")        or "No especificado",
                     exp.get("fecha_inicio") or "No especificada",
                     fecha_fin,
                     estado]

            fonts2 = [DATA_FONT, DATA_FONT, DATA_FONT, DATA_FONT, DATA_FONT, f_fin, f_est]

            for ci,(v,fnt) in enumerate(zip(vals2,fonts2),1):
                aln = AC if ci in (1,7) else AL
                _celda(ws2,fila_exp,ci,v, font=fnt, fill=fill, align=aln, border=BRD2)
            ws2.row_dimensions[fila_exp].height = 22
            fila_exp+=1

    total_exp = fila_exp - 2
    c2 = ws2.cell(row=fila_exp, column=1,
                  value=f"Total: {total_exp} registros de experiencia laboral")
    c2.font  = Font(name="Calibri",bold=True,color="86EFAC",size=10)
    c2.fill  = PatternFill("solid",fgColor="0A1A0A")
    ws2.merge_cells(f"A{fila_exp}:{get_column_letter(len(cols2))}{fila_exp}")
    c2.alignment = AC
    ws2.freeze_panes = "A2"

    wb.save(ruta)


# ─── APP ──────────────────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Analizador de HV · Groq IA")
        self.geometry("1280x840"); self.minsize(1050,700)
        self.configure(fg_color=C["bg"])
        self.archivos=[]; self.resultados=[]; self.procesando=False
        self._ui()

    def _ui(self):
        # Sidebar
        sb = ctk.CTkFrame(self, width=268, corner_radius=0, fg_color=C["card"])
        sb.pack(side="left", fill="y"); sb.pack_propagate(False)

        ctk.CTkLabel(sb, text="⚡", font=ctk.CTkFont(size=36),
                     text_color=C["ac"]).pack(pady=(22,0))
        ctk.CTkLabel(sb, text="Analizador de HV",
                     font=ctk.CTkFont(size=17,weight="bold"),
                     text_color=C["txt"]).pack(pady=(4,2))
        ctk.CTkLabel(sb, text="Powered by Groq · Llama 3.3",
                     font=ctk.CTkFont(size=10), text_color=C["ac2"]).pack(pady=(0,14))
        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        ctk.CTkLabel(sb, text="🔑  API Key de Groq",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(anchor="w", padx=20, pady=(14,4))
        ctk.CTkLabel(sb, text="Gratis en: console.groq.com",
                     font=ctk.CTkFont(size=9), text_color=C["ok"]).pack(anchor="w", padx=22, pady=(0,6))
        self.entry_api = ctk.CTkEntry(sb, placeholder_text="gsk_...",
                                      show="•", height=36, corner_radius=8,
                                      fg_color=C["input"], border_color=C["borde"])
        self.entry_api.pack(fill="x", padx=20)
        ctk.CTkButton(sb, text="👁  Mostrar / Ocultar", height=26,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], font=ctk.CTkFont(size=10), hover_color=C["input"],
                      command=lambda: self.entry_api.configure(
                          show="" if self.entry_api.cget("show")=="•" else "•")
                      ).pack(fill="x", padx=20, pady=(4,14))

        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        ctk.CTkLabel(sb, text="📂  Archivos CV",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(anchor="w", padx=20, pady=(14,8))
        ctk.CTkButton(sb, text="＋  Agregar HV", height=38, corner_radius=8,
                      fg_color=C["ac"], hover_color="#EA580C",
                      font=ctk.CTkFont(size=12,weight="bold"),
                      command=self._agregar).pack(fill="x", padx=20)
        ctk.CTkButton(sb, text="🗑  Limpiar todo", height=30, corner_radius=8,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], hover_color=C["input"],
                      command=self._limpiar).pack(fill="x", padx=20, pady=(6,0))
        self.lbl_n = ctk.CTkLabel(sb, text="0 archivos cargados",
                                   font=ctk.CTkFont(size=10), text_color=C["dim"])
        self.lbl_n.pack(pady=6)

        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        self.btn_ana = ctk.CTkButton(sb, text="⚡  ANALIZAR", height=46, corner_radius=10,
                                      fg_color=C["ok"], hover_color="#059669",
                                      font=ctk.CTkFont(size=14,weight="bold"),
                                      command=self._iniciar)
        self.btn_ana.pack(fill="x", padx=20, pady=(14,6))

        self.btn_exp = ctk.CTkButton(sb, text="📊  Exportar Excel", height=38, corner_radius=10,
                                      fg_color="#7C3AED", hover_color="#6D28D9",
                                      font=ctk.CTkFont(size=12,weight="bold"),
                                      state="disabled", command=self._exportar)
        self.btn_exp.pack(fill="x", padx=20)

        # Info pestañas Excel
        info = ctk.CTkFrame(sb, fg_color="#1A1A35", corner_radius=8)
        info.pack(fill="x", padx=20, pady=(10,0))
        ctk.CTkLabel(info, text="El Excel tendrá 2 pestañas:",
                     font=ctk.CTkFont(size=9,weight="bold"),
                     text_color=C["ac2"]).pack(anchor="w", padx=10, pady=(8,2))
        ctk.CTkLabel(info, text="📋  Datos Personales",
                     font=ctk.CTkFont(size=9), text_color=C["txt"]).pack(anchor="w", padx=14, pady=1)
        ctk.CTkLabel(info, text="💼  Experiencia Laboral",
                     font=ctk.CTkFont(size=9), text_color=C["lok"]).pack(anchor="w", padx=14, pady=(1,8))

        ctk.CTkLabel(sb, text="", fg_color="transparent").pack(expand=True)
        ctk.CTkLabel(sb, text=f"Modelo: {MODELO}",
                     font=ctk.CTkFont(size=9), text_color=C["dim"]).pack(pady=(0,4))
        ctk.CTkLabel(sb, text="PDF · DOCX · TXT",
                     font=ctk.CTkFont(size=9), text_color=C["dim"]).pack(pady=(0,14))

        # Panel derecho con tabs
        panel = ctk.CTkFrame(self, fg_color=C["bg"])
        panel.pack(side="right", fill="both", expand=True, padx=14, pady=14)

        # Header
        hdr = ctk.CTkFrame(panel, fg_color="transparent")
        hdr.pack(fill="x", pady=(0,6))
        ctk.CTkLabel(hdr, text="Resultados",
                     font=ctk.CTkFont(size=17,weight="bold"),
                     text_color=C["txt"]).pack(side="left")
        self.lbl_est = ctk.CTkLabel(hdr, text="●  Listo",
                                     font=ctk.CTkFont(size=11), text_color=C["ok"])
        self.lbl_est.pack(side="right")

        self.prog = ctk.CTkProgressBar(panel, height=5,
                                        fg_color=C["card"], progress_color=C["ac"])
        self.prog.pack(fill="x"); self.prog.set(0)
        self.lbl_prog = ctk.CTkLabel(panel, text="",
                                      font=ctk.CTkFont(size=9), text_color=C["dim"])
        self.lbl_prog.pack(anchor="e", pady=(2,4))

        # Tabs
        self.tabs = ctk.CTkTabview(panel, fg_color=C["card"],
                                    segmented_button_fg_color=C["input"],
                                    segmented_button_selected_color=C["ac"],
                                    segmented_button_selected_hover_color="#EA580C")
        self.tabs.pack(fill="both", expand=True)
        self.tabs.add("📋  Datos Personales")
        self.tabs.add("💼  Experiencia Laboral")

        # Tab 1: Datos personales
        self._tabla_personal(self.tabs.tab("📋  Datos Personales"))
        # Tab 2: Experiencia
        self._tabla_exp(self.tabs.tab("💼  Experiencia Laboral"))

        # Consola
        ctk.CTkFrame(panel, height=1, fg_color=C["borde"]).pack(fill="x", pady=(6,0))
        cab_log = ctk.CTkFrame(panel, fg_color="transparent", height=28)
        cab_log.pack(fill="x", pady=(4,2)); cab_log.pack_propagate(False)
        ctk.CTkLabel(cab_log, text="🖥  Consola",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(side="left")
        ctk.CTkButton(cab_log, text="Limpiar", height=22, width=70,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], font=ctk.CTkFont(size=9), hover_color=C["input"],
                      command=self._limpiar_consola).pack(side="right")
        self.consola = ctk.CTkTextbox(panel, height=140, corner_radius=8,
                                       fg_color="#080812", text_color=C["txt"],
                                       font=ctk.CTkFont(size=10, family="Courier"))
        self.consola.pack(fill="x", pady=(0,4))
        self.consola.tag_config("ok",   foreground=C["lok"])
        self.consola.tag_config("err",  foreground=C["lerr"])
        self.consola.tag_config("info", foreground=C["linfo"])
        self.consola.tag_config("warn", foreground=C["lwarn"])
        self.consola.tag_config("dim",  foreground=C["dim"])
        self._log("info","Sistema listo. Agrega HV y presiona ⚡ ANALIZAR.")
        self._log("info",f"Extrae: datos personales + experiencia laboral completa")

    # ── Tabla datos personales ────────────────────────────────────────────────
    def _tabla_personal(self, parent):
        COLS   = ["#","Nombre","Correo","Celular","Universidad","Nivel","Título","Exps."]
        ANCHOS = [35,170,180,100,155,90,150,55]
        cab = ctk.CTkFrame(parent, fg_color="#431407", corner_radius=0, height=34)
        cab.pack(fill="x", padx=1, pady=(4,0)); cab.pack_propagate(False)
        for col,ancho in zip(COLS,ANCHOS):
            ctk.CTkLabel(cab, text=col, font=ctk.CTkFont(size=10,weight="bold"),
                         text_color=C["ac2"], width=ancho, anchor="w"
                         ).pack(side="left", padx=(8,0))
        self.filas_p = ctk.CTkScrollableFrame(parent, fg_color=C["card"],
                                               scrollbar_button_color=C["ac"], corner_radius=0)
        self.filas_p.pack(fill="both", expand=True, padx=1, pady=(0,1))
        self.COLS_P=COLS; self.ANCHOS_P=ANCHOS

    # ── Tabla experiencia ─────────────────────────────────────────────────────
    def _tabla_exp(self, parent):
        COLS   = ["#","Nombre","Empresa","Cargo","Fecha Inicio","Fecha Fin","Estado"]
        ANCHOS = [35,155,170,155,100,100,90]
        cab = ctk.CTkFrame(parent, fg_color="#1A3A1A", corner_radius=0, height=34)
        cab.pack(fill="x", padx=1, pady=(4,0)); cab.pack_propagate(False)
        for col,ancho in zip(COLS,ANCHOS):
            ctk.CTkLabel(cab, text=col, font=ctk.CTkFont(size=10,weight="bold"),
                         text_color="#86EFAC", width=ancho, anchor="w"
                         ).pack(side="left", padx=(8,0))
        self.filas_e = ctk.CTkScrollableFrame(parent, fg_color=C["card"],
                                               scrollbar_button_color="#16A34A", corner_radius=0)
        self.filas_e.pack(fill="both", expand=True, padx=1, pady=(0,1))
        self.COLS_E=COLS; self.ANCHOS_E=ANCHOS

    # ── Consola ───────────────────────────────────────────────────────────────
    def _log(self, nivel, texto):
        ts  = datetime.now().strftime("%H:%M:%S")
        ico = {"ok":"✓","err":"✗","info":"›","warn":"⚠"}.get(nivel,"·")
        def _ins():
            self.consola.configure(state="normal")
            self.consola.insert("end", f"[{ts}] ","dim")
            self.consola.insert("end", f"{ico} {texto}\n", nivel)
            self.consola.see("end")
            self.consola.configure(state="disabled")
        self.after(0,_ins)

    def _limpiar_consola(self):
        self.consola.configure(state="normal")
        self.consola.delete("1.0","end")
        self.consola.configure(state="disabled")
        self._log("info","Consola limpiada.")

    # ── Archivos ──────────────────────────────────────────────────────────────
    def _agregar(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar Hojas de Vida",
            filetypes=[("CV","*.pdf *.docx *.doc *.txt"),
                       ("PDF","*.pdf"),("Word","*.docx *.doc"),("Texto","*.txt")])
        nuevos=[r for r in rutas if r not in self.archivos]
        self.archivos.extend(nuevos)
        if nuevos:
            self._log("info","Agregados: "+", ".join(Path(r).name for r in nuevos))
        self.lbl_n.configure(text=f"{len(self.archivos)} archivo(s) cargado(s)")

    def _limpiar(self):
        self.archivos.clear(); self.resultados.clear()
        self.lbl_n.configure(text="0 archivos cargados")
        for w in self.filas_p.winfo_children(): w.destroy()
        for w in self.filas_e.winfo_children(): w.destroy()
        self.btn_exp.configure(state="disabled")
        self.prog.set(0); self.lbl_prog.configure(text="")
        self._log("warn","Lista limpiada.")

    def _agregar_fila_personal(self, idx, datos):
        bg  = C["input"] if idx%2==0 else C["card"]
        err = "ERROR" in str(datos.get("nombre_completo",""))
        f   = ctk.CTkFrame(self.filas_p, fg_color=bg, height=30, corner_radius=0)
        f.pack(fill="x", pady=1); f.pack_propagate(False)
        exp_count = len(datos.get("experiencia_laboral") or [])
        vals=[str(idx),
              datos.get("nombre_completo") or "—",
              datos.get("correo")          or "—",
              datos.get("celular")         or "—",
              datos.get("universidad")     or "—",
              datos.get("nivel_estudios")  or "—",
              datos.get("titulo")          or "—",
              str(exp_count)]
        for i,(v,ancho) in enumerate(zip(vals,self.ANCHOS_P)):
            col = C["lerr"] if err else (C["ac2"] if i==0 else (C["lok"] if i==7 else C["txt"]))
            ctk.CTkLabel(f, text=v[:22]+("…" if len(v)>22 else ""),
                         font=ctk.CTkFont(size=9), text_color=col,
                         width=ancho, anchor="w").pack(side="left", padx=(8,0))

    def _agregar_filas_exp(self, idx_persona, datos):
        nombre = datos.get("nombre_completo") or "—"
        exps   = datos.get("experiencia_laboral") or []
        if not exps:
            return
        for exp in exps:
            fecha_fin_raw = str(exp.get("fecha_fin","")).strip().lower()
            en_curso = fecha_fin_raw in ("en curso","actual","presente","current",
                                          "actualidad","a la fecha","","none","null")
            fecha_fin = "En curso" if en_curso else (exp.get("fecha_fin") or "N/D")
            estado    = "🟢 En curso" if en_curso else "✓ Fin"
            bg = "#14350C" if en_curso else (C["input"] if len(self.filas_e.winfo_children())%2==0 else C["card"])

            f = ctk.CTkFrame(self.filas_e, fg_color=bg, height=30, corner_radius=0)
            f.pack(fill="x", pady=1); f.pack_propagate(False)
            vals=[str(idx_persona), nombre,
                  exp.get("empresa")      or "—",
                  exp.get("cargo")        or "—",
                  exp.get("fecha_inicio") or "—",
                  fecha_fin, estado]
            for i,(v,ancho) in enumerate(zip(vals,self.ANCHOS_E)):
                col = ("#4ADE80" if en_curso else C["txt"]) if i>1 else C["ac2"]
                if i==6: col = "#4ADE80" if en_curso else C["dim"]
                ctk.CTkLabel(f, text=v[:22]+("…" if len(v)>22 else ""),
                             font=ctk.CTkFont(size=9), text_color=col,
                             width=ancho, anchor="w").pack(side="left", padx=(8,0))

    # ── Análisis ──────────────────────────────────────────────────────────────
    def _iniciar(self):
        if self.procesando: return
        api_key=self.entry_api.get().strip()
        if not api_key:
            self._log("err","API Key vacía.")
            messagebox.showerror("API Key requerida","Obtén tu clave gratis en:\nconsole.groq.com")
            return
        if not self.archivos:
            self._log("warn","No hay archivos cargados.")
            messagebox.showwarning("Sin archivos","Agrega al menos una hoja de vida.")
            return
        self.procesando=True; self.resultados.clear()
        for w in self.filas_p.winfo_children(): w.destroy()
        for w in self.filas_e.winfo_children(): w.destroy()
        self.btn_ana.configure(state="disabled", text="⏳  Analizando…")
        self.btn_exp.configure(state="disabled")
        self._log("info",f"Iniciando análisis de {len(self.archivos)} archivo(s)…")
        threading.Thread(target=self._thread, args=(api_key,), daemon=True).start()

    def _thread(self, api_key):
        total=len(self.archivos); errores=0
        for i,ruta in enumerate(self.archivos):
            nombre=Path(ruta).name
            self._set_est(f"⏳  {nombre}", C["warn"])
            self._set_prog(i/total, f"Procesando {i+1}/{total}: {nombre}")
            self._log("info",f"[{i+1}/{total}] Leyendo: {nombre}")
            try:
                texto=extraer_texto(ruta)
                if not texto.strip():
                    raise ValueError("No se pudo extraer texto del archivo")
                self._log("ok",f"  Texto extraído: {len(texto)} caracteres")
                self._log("info","  Analizando con Groq…")
                datos=analizar_con_groq(texto, api_key)
                datos["_archivo"]=nombre
                self.resultados.append(datos)
                exp_count=len(datos.get("experiencia_laboral") or [])
                self._log("ok",
                    f"  ✓ {datos.get('nombre_completo') or 'Sin nombre'} "
                    f"· {exp_count} experiencia(s) laboral(es)")
                idx=len(self.resultados)
                self.after(0, lambda d=datos, ix=idx: (
                    self._agregar_fila_personal(ix, d),
                    self._agregar_filas_exp(ix, d)
                ))
            except Exception as e:
                errores+=1
                tb=traceback.format_exc()
                self._log("err",f"  {type(e).__name__}: {e}")
                for linea in tb.strip().split("\n")[-5:]:
                    self._log("err",f"    {linea}")
                de={"_archivo":nombre,
                    "nombre_completo":f"ERROR: {str(e)[:55]}",
                    "correo":"—","celular":"—","universidad":"—",
                    "nivel_estudios":"—","titulo":"—","experiencia_laboral":[]}
                self.resultados.append(de)
                self.after(0, lambda d=de, ix=len(self.resultados):
                           self._agregar_fila_personal(ix, d))

        msg=(f"Completado: {len(self.resultados)} HV"
             +(f", {errores} error(es)" if errores else " ✓ sin errores"))
        self._log("ok" if not errores else "warn", msg)
        self._set_prog(1.0, msg)
        self._set_est("●  Completado", C["ok"] if not errores else C["warn"])
        self.after(0, self._fin)

    def _fin(self):
        self.procesando=False
        self.btn_ana.configure(state="normal", text="⚡  ANALIZAR")
        if self.resultados: self.btn_exp.configure(state="normal")

    def _set_est(self,t,col): self.after(0,lambda:self.lbl_est.configure(text=t,text_color=col))
    def _set_prog(self,v,t):
        self.after(0,lambda:self.prog.set(v))
        self.after(0,lambda:self.lbl_prog.configure(text=t))

    def _exportar(self):
        if not self.resultados: return
        fecha=datetime.now().strftime("%Y%m%d_%H%M")
        ruta=filedialog.asksaveasfilename(
            title="Guardar Excel", defaultextension=".xlsx",
            initialfile=f"hojas_de_vida_{fecha}.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not ruta: return
        try:
            exportar_excel(self.resultados, ruta)
            total_exp=sum(len(d.get("experiencia_laboral") or []) for d in self.resultados)
            self._log("ok",f"Excel exportado: {ruta}")
            messagebox.showinfo("✅ Excel exportado",
                f"Archivo guardado exitosamente.\n\n"
                f"📋 Pestaña 1 — Datos Personales: {len(self.resultados)} personas\n"
                f"💼 Pestaña 2 — Experiencia Laboral: {total_exp} registros\n\n"
                f"{ruta}")
        except Exception as e:
            self._log("err",f"Error: {e}")
            messagebox.showerror("Error al exportar", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
