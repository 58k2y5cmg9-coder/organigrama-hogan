"""
Analizador de Hojas de Vida — powered by Groq (gratis, sin tarjeta)
Extrae: Nombre, Correo, Celular, Universidad, Estudios
Exporta a Excel con columnas separadas
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading, json, traceback
import pdfplumber, docx, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from groq import Groq

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

C = {
    "bg":    "#0D0D1A", "card":  "#141428", "input": "#1A1A35",
    "ac":    "#F97316", "ac2":   "#FED7AA", "ok":    "#10B981",
    "err":   "#EF4444", "warn":  "#F59E0B", "txt":   "#E2E8F0",
    "dim":   "#94A3B8", "borde": "#2D2D5E",
    "lok":   "#34D399", "lerr":  "#F87171", "linfo": "#60A5FA", "lwarn": "#FBBF24",
}
MODELO = "llama-3.3-70b-versatile"


# ─── EXTRACCIÓN DE TEXTO ──────────────────────────────────────────────────────
def extraer_texto(ruta):
    ext = Path(ruta).suffix.lower()
    if ext == ".pdf":
        texto = ""
        with pdfplumber.open(ruta) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t: texto += t + "\n"
        return texto
    elif ext in (".docx", ".doc"):
        d = docx.Document(ruta)
        return "\n".join(p.text for p in d.paragraphs if p.text.strip())
    elif ext == ".txt":
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    return ""


# ─── ANÁLISIS CON GROQ ────────────────────────────────────────────────────────
def analizar_con_groq(texto_cv, api_key):
    cliente = Groq(api_key=api_key)
    prompt = (
        "Analiza esta hoja de vida y extrae la información solicitada.\n"
        "Responde ÚNICAMENTE con un objeto JSON válido, sin markdown, sin explicaciones.\n"
        "Si no encuentras un dato usa null.\n\n"
        "Hoja de vida:\n---\n"
        + texto_cv[:6000]
        + "\n---\n\n"
        "JSON a retornar:\n"
        '{"nombre_completo":null,"correo":null,"celular":null,'
        '"universidad":null,"nivel_estudios":null,"titulo":null,"anio_graduacion":null}'
    )
    respuesta = cliente.chat.completions.create(
        model=MODELO,
        messages=[
            {"role": "system", "content": "Eres un extractor de datos de hojas de vida. "
             "Solo respondes con JSON válido, sin texto adicional ni markdown."},
            {"role": "user", "content": prompt}
        ],
        temperature=0,
        max_tokens=500,
    )
    texto = respuesta.choices[0].message.content.strip()
    if "```" in texto:
        for parte in texto.split("```"):
            limpio = parte.strip().lstrip("json").strip()
            if limpio.startswith("{"): texto = limpio; break
    return json.loads(texto)


# ─── EXPORTAR EXCEL ───────────────────────────────────────────────────────────
def exportar_excel(datos, ruta):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Hojas de Vida"
    fh = PatternFill("solid", fgColor="431407")
    fa = PatternFill("solid", fgColor="1C1917")
    fn = PatternFill("solid", fgColor="0C0A09")
    brd = Border(
        left=Side(style="thin", color="7C2D12"),
        right=Side(style="thin", color="7C2D12"),
        top=Side(style="thin", color="7C2D12"),
        bottom=Side(style="thin", color="7C2D12"),
    )
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    cols = [("N°",8),("Nombre Completo",30),("Correo",32),("Celular",18),
            ("Universidad",30),("Nivel Estudios",18),("Título / Carrera",28),
            ("Año Graduación",16),("Archivo Origen",35)]
    for ci,(titulo,ancho) in enumerate(cols,1):
        c = ws.cell(row=1,column=ci,value=titulo)
        c.font = Font(name="Calibri",bold=True,color="FED7AA",size=11)
        c.fill=fh; c.alignment=ac; c.border=brd
        ws.column_dimensions[get_column_letter(ci)].width=ancho
    ws.row_dimensions[1].height=28
    for i,fila in enumerate(datos,1):
        fill = fa if i%2==0 else fn
        vals=[i,
              fila.get("nombre_completo") or "No encontrado",
              fila.get("correo")          or "No encontrado",
              fila.get("celular")         or "No encontrado",
              fila.get("universidad")     or "No encontrado",
              fila.get("nivel_estudios")  or "No encontrado",
              fila.get("titulo")          or "No encontrado",
              fila.get("anio_graduacion") or "No encontrado",
              fila.get("_archivo","")]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=i+1,column=ci,value=v)
            c.fill=fill; c.border=brd
            c.alignment=ac if ci==1 else al
            c.font=Font(name="Calibri",
                        color="94A3B8" if ci==len(vals) else "E2E8F0",
                        size=9 if ci==len(vals) else 10,
                        italic=(ci==len(vals)))
        ws.row_dimensions[i+1].height=22
    ft=len(datos)+2
    c=ws.cell(row=ft,column=1,value=f"Total: {len(datos)} registros")
    c.font=Font(name="Calibri",bold=True,color="FED7AA",size=10)
    c.fill=PatternFill("solid",fgColor="1C0A00")
    ws.merge_cells(f"A{ft}:I{ft}"); c.alignment=ac
    ws.freeze_panes="A2"; wb.save(ruta)


# ─── APP ──────────────────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Analizador de HV · Groq IA")
        self.geometry("1200x800"); self.minsize(1000,680)
        self.configure(fg_color=C["bg"])
        self.archivos=[]; self.resultados=[]; self.procesando=False
        self._ui()

    def _ui(self):
        # Sidebar
        sb = ctk.CTkFrame(self, width=264, corner_radius=0, fg_color=C["card"])
        sb.pack(side="left", fill="y"); sb.pack_propagate(False)

        ctk.CTkLabel(sb, text="⚡", font=ctk.CTkFont(size=36),
                     text_color=C["ac"]).pack(pady=(22,0))
        ctk.CTkLabel(sb, text="Analizador de HV",
                     font=ctk.CTkFont(size=17,weight="bold"),
                     text_color=C["txt"]).pack(pady=(4,2))
        ctk.CTkLabel(sb, text="Powered by Groq · Llama 3.3",
                     font=ctk.CTkFont(size=10), text_color=C["ac2"]).pack(pady=(0,14))
        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        ctk.CTkLabel(sb, text="🔑  API Key de Groq",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(anchor="w", padx=20, pady=(14,4))
        ctk.CTkLabel(sb, text="Gratis en: console.groq.com",
                     font=ctk.CTkFont(size=9), text_color=C["ok"]).pack(anchor="w", padx=22, pady=(0,6))

        self.entry_api = ctk.CTkEntry(sb, placeholder_text="gsk_...",
                                      show="•", height=36, corner_radius=8,
                                      fg_color=C["input"], border_color=C["borde"])
        self.entry_api.pack(fill="x", padx=20)
        ctk.CTkButton(sb, text="👁  Mostrar / Ocultar", height=26,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], font=ctk.CTkFont(size=10),
                      hover_color=C["input"],
                      command=lambda: self.entry_api.configure(
                          show="" if self.entry_api.cget("show")=="•" else "•")
                      ).pack(fill="x", padx=20, pady=(4,14))

        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        ctk.CTkLabel(sb, text="📂  Archivos CV",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(anchor="w", padx=20, pady=(14,8))
        ctk.CTkButton(sb, text="＋  Agregar HV", height=38, corner_radius=8,
                      fg_color=C["ac"], hover_color="#EA580C",
                      font=ctk.CTkFont(size=12,weight="bold"),
                      command=self._agregar).pack(fill="x", padx=20)
        ctk.CTkButton(sb, text="🗑  Limpiar todo", height=30, corner_radius=8,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], hover_color=C["input"],
                      command=self._limpiar).pack(fill="x", padx=20, pady=(6,0))
        self.lbl_n = ctk.CTkLabel(sb, text="0 archivos cargados",
                                   font=ctk.CTkFont(size=10), text_color=C["dim"])
        self.lbl_n.pack(pady=6)

        ctk.CTkFrame(sb, height=1, fg_color=C["borde"]).pack(fill="x", padx=20)

        self.btn_ana = ctk.CTkButton(sb, text="⚡  ANALIZAR", height=46,
                                      corner_radius=10, fg_color=C["ok"],
                                      hover_color="#059669",
                                      font=ctk.CTkFont(size=14,weight="bold"),
                                      command=self._iniciar)
        self.btn_ana.pack(fill="x", padx=20, pady=(14,6))

        self.btn_exp = ctk.CTkButton(sb, text="📊  Exportar Excel", height=38,
                                      corner_radius=10, fg_color="#7C3AED",
                                      hover_color="#6D28D9",
                                      font=ctk.CTkFont(size=12,weight="bold"),
                                      state="disabled", command=self._exportar)
        self.btn_exp.pack(fill="x", padx=20)

        ctk.CTkLabel(sb, text="", fg_color="transparent").pack(expand=True)
        ctk.CTkLabel(sb, text=f"Modelo: {MODELO}",
                     font=ctk.CTkFont(size=9), text_color=C["dim"]).pack(pady=(0,4))
        ctk.CTkLabel(sb, text="PDF · DOCX · TXT",
                     font=ctk.CTkFont(size=9), text_color=C["dim"]).pack(pady=(0,14))

        # Panel derecho
        panel = ctk.CTkFrame(self, fg_color=C["bg"])
        panel.pack(side="right", fill="both", expand=True, padx=14, pady=14)

        hdr = ctk.CTkFrame(panel, fg_color="transparent")
        hdr.pack(fill="x", pady=(0,8))
        ctk.CTkLabel(hdr, text="Resultados",
                     font=ctk.CTkFont(size=17,weight="bold"),
                     text_color=C["txt"]).pack(side="left")
        self.lbl_est = ctk.CTkLabel(hdr, text="●  Listo",
                                     font=ctk.CTkFont(size=11), text_color=C["ok"])
        self.lbl_est.pack(side="right")

        self.prog = ctk.CTkProgressBar(panel, height=5,
                                        fg_color=C["card"], progress_color=C["ac"])
        self.prog.pack(fill="x"); self.prog.set(0)
        self.lbl_prog = ctk.CTkLabel(panel, text="",
                                      font=ctk.CTkFont(size=9), text_color=C["dim"])
        self.lbl_prog.pack(anchor="e", pady=(2,6))

        # Tabla
        cont = ctk.CTkFrame(panel, fg_color=C["card"], corner_radius=10)
        cont.pack(fill="both", expand=True)
        COLS   = ["#","Nombre","Correo","Celular","Universidad","Nivel","Título"]
        ANCHOS = [38,175,185,105,165,95,155]
        cab = ctk.CTkFrame(cont, fg_color="#431407", corner_radius=0, height=34)
        cab.pack(fill="x", padx=1, pady=(1,0)); cab.pack_propagate(False)
        for col,ancho in zip(COLS,ANCHOS):
            ctk.CTkLabel(cab, text=col, font=ctk.CTkFont(size=10,weight="bold"),
                         text_color=C["ac2"], width=ancho, anchor="w"
                         ).pack(side="left", padx=(8,0))
        self.filas = ctk.CTkScrollableFrame(cont, fg_color=C["card"],
                                             scrollbar_button_color=C["ac"],
                                             corner_radius=0)
        self.filas.pack(fill="both", expand=True, padx=1, pady=(0,1))
        self.COLS=COLS; self.ANCHOS=ANCHOS

        # Consola
        ctk.CTkFrame(panel, height=1, fg_color=C["borde"]).pack(fill="x", pady=(8,0))
        cab_log = ctk.CTkFrame(panel, fg_color="transparent", height=28)
        cab_log.pack(fill="x", pady=(6,2)); cab_log.pack_propagate(False)
        ctk.CTkLabel(cab_log, text="🖥  Consola de eventos y errores",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["ac2"]).pack(side="left")
        ctk.CTkButton(cab_log, text="Limpiar", height=22, width=80,
                      fg_color="transparent", border_width=1, border_color=C["borde"],
                      text_color=C["dim"], font=ctk.CTkFont(size=9),
                      hover_color=C["input"],
                      command=self._limpiar_consola).pack(side="right")
        self.consola = ctk.CTkTextbox(panel, height=155, corner_radius=8,
                                       fg_color="#080812", text_color=C["txt"],
                                       font=ctk.CTkFont(size=10, family="Courier"))
        self.consola.pack(fill="x", pady=(0,4))
        self.consola.tag_config("ok",   foreground=C["lok"])
        self.consola.tag_config("err",  foreground=C["lerr"])
        self.consola.tag_config("info", foreground=C["linfo"])
        self.consola.tag_config("warn", foreground=C["lwarn"])
        self.consola.tag_config("dim",  foreground=C["dim"])

        self._log("info", "Sistema listo. Agrega HV y presiona ⚡ ANALIZAR.")
        self._log("info", f"Modelo: {MODELO}  |  API Key gratis en console.groq.com")

    # ── Consola ───────────────────────────────────────────────────────────────
    def _log(self, nivel, texto):
        ts  = datetime.now().strftime("%H:%M:%S")
        ico = {"ok":"✓","err":"✗","info":"›","warn":"⚠"}.get(nivel,"·")
        def _ins():
            self.consola.configure(state="normal")
            self.consola.insert("end", f"[{ts}] ", "dim")
            self.consola.insert("end", f"{ico} {texto}\n", nivel)
            self.consola.see("end")
            self.consola.configure(state="disabled")
        self.after(0, _ins)

    def _limpiar_consola(self):
        self.consola.configure(state="normal")
        self.consola.delete("1.0","end")
        self.consola.configure(state="disabled")
        self._log("info","Consola limpiada.")

    # ── Archivos ──────────────────────────────────────────────────────────────
    def _agregar(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar Hojas de Vida",
            filetypes=[("CV","*.pdf *.docx *.doc *.txt"),
                       ("PDF","*.pdf"),("Word","*.docx *.doc"),("Texto","*.txt")])
        nuevos = [r for r in rutas if r not in self.archivos]
        self.archivos.extend(nuevos)
        if nuevos:
            self._log("info","Agregados: "+", ".join(Path(r).name for r in nuevos))
        self.lbl_n.configure(text=f"{len(self.archivos)} archivo(s) cargado(s)")

    def _limpiar(self):
        self.archivos.clear(); self.resultados.clear()
        self.lbl_n.configure(text="0 archivos cargados")
        for w in self.filas.winfo_children(): w.destroy()
        self.btn_exp.configure(state="disabled")
        self.prog.set(0); self.lbl_prog.configure(text="")
        self._log("warn","Lista limpiada.")

    def _fila_tabla(self, idx, datos):
        bg  = C["input"] if idx%2==0 else C["card"]
        err = "ERROR" in str(datos.get("nombre_completo",""))
        f   = ctk.CTkFrame(self.filas, fg_color=bg, height=30, corner_radius=0)
        f.pack(fill="x", pady=1); f.pack_propagate(False)
        vals=[str(idx),
              datos.get("nombre_completo") or "—",
              datos.get("correo")          or "—",
              datos.get("celular")         or "—",
              datos.get("universidad")     or "—",
              datos.get("nivel_estudios")  or "—",
              datos.get("titulo")          or "—"]
        for i,(v,ancho) in enumerate(zip(vals,self.ANCHOS)):
            col = C["lerr"] if err else (C["ac2"] if i==0 else C["txt"])
            ctk.CTkLabel(f, text=v[:24]+("…" if len(v)>24 else ""),
                         font=ctk.CTkFont(size=9), text_color=col,
                         width=ancho, anchor="w").pack(side="left", padx=(8,0))

    # ── Análisis ──────────────────────────────────────────────────────────────
    def _iniciar(self):
        if self.procesando: return
        api_key = self.entry_api.get().strip()
        if not api_key:
            self._log("err","API Key vacía.")
            messagebox.showerror("API Key requerida",
                                 "Obtén tu clave gratis en:\nconsole.groq.com → API Keys")
            return
        if not self.archivos:
            self._log("warn","No hay archivos cargados.")
            messagebox.showwarning("Sin archivos","Agrega al menos una hoja de vida.")
            return
        self.procesando=True; self.resultados.clear()
        for w in self.filas.winfo_children(): w.destroy()
        self.btn_ana.configure(state="disabled", text="⏳  Analizando…")
        self.btn_exp.configure(state="disabled")
        self._log("info",f"Iniciando análisis de {len(self.archivos)} archivo(s)…")
        threading.Thread(target=self._thread, args=(api_key,), daemon=True).start()

    def _thread(self, api_key):
        total=len(self.archivos); errores=0
        for i,ruta in enumerate(self.archivos):
            nombre=Path(ruta).name
            self._set_est(f"⏳  {nombre}", C["warn"])
            self._set_prog(i/total, f"Procesando {i+1}/{total}: {nombre}")
            self._log("info",f"[{i+1}/{total}] Leyendo: {nombre}")
            try:
                texto=extraer_texto(ruta)
                if not texto.strip():
                    raise ValueError("No se pudo extraer texto (¿PDF escaneado?)")
                self._log("ok",f"  Texto extraído: {len(texto)} caracteres")
                self._log("info","  Enviando a Groq…")
                datos=analizar_con_groq(texto, api_key)
                datos["_archivo"]=nombre
                self.resultados.append(datos)
                self._log("ok",f"  ✓ {datos.get('nombre_completo') or 'nombre no encontrado'}")
                self.after(0, lambda d=datos, idx=len(self.resultados):
                           self._fila_tabla(idx,d))
            except Exception as e:
                errores+=1
                tb=traceback.format_exc()
                self._log("err",f"  {type(e).__name__}: {e}")
                for linea in tb.strip().split("\n")[-5:]:
                    self._log("err",f"    {linea}")
                de={"_archivo":nombre,
                    "nombre_completo":f"ERROR: {type(e).__name__}: {str(e)[:60]}",
                    "correo":"—","celular":"—","universidad":"—",
                    "nivel_estudios":"—","titulo":"—"}
                self.resultados.append(de)
                self.after(0, lambda d=de, idx=len(self.resultados):
                           self._fila_tabla(idx,d))

        msg=(f"Completado: {len(self.resultados)} HV"
             +(f", {errores} errores" if errores else " ✓ sin errores"))
        self._log("ok" if not errores else "warn", msg)
        self._set_prog(1.0, msg)
        self._set_est("●  Completado", C["ok"] if not errores else C["warn"])
        self.after(0, self._fin)

    def _fin(self):
        self.procesando=False
        self.btn_ana.configure(state="normal", text="⚡  ANALIZAR")
        if self.resultados: self.btn_exp.configure(state="normal")

    def _set_est(self,t,col): self.after(0,lambda:self.lbl_est.configure(text=t,text_color=col))
    def _set_prog(self,v,t):
        self.after(0,lambda:self.prog.set(v))
        self.after(0,lambda:self.lbl_prog.configure(text=t))

    def _exportar(self):
        if not self.resultados: return
        fecha=datetime.now().strftime("%Y%m%d_%H%M")
        ruta=filedialog.asksaveasfilename(
            title="Guardar Excel", defaultextension=".xlsx",
            initialfile=f"hojas_de_vida_{fecha}.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not ruta: return
        try:
            exportar_excel(self.resultados, ruta)
            self._log("ok",f"Excel exportado: {ruta}")
            messagebox.showinfo("✅ Listo",
                                f"Excel guardado:\n{ruta}\n\n{len(self.resultados)} registros.")
        except Exception as e:
            self._log("err",f"Error: {e}")
            messagebox.showerror("Error al exportar", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
